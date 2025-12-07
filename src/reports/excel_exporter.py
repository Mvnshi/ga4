"""
Excel Exporter
==============
Professional Excel report generation with formatting.

Creates multi-sheet workbooks with:
- Executive summary
- Detailed data tables
- Conditional formatting
- Charts
"""

from pathlib import Path
from typing import Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

from config.settings import OUTPUT_DIR, ClientConfig


class ExcelExporter:
    """
    Creates professional Excel reports.
    
    Features:
    - Multiple sheets organized by topic
    - Executive summary sheet
    - Conditional formatting for trends
    - Professional styling
    """
    
    # Color palette
    COLORS = {
        'primary': '2D5016',      # Forest green
        'secondary': 'F4C430',    # Honey gold
        'header': '1F4E0F',       # Dark green
        'positive': '22C55E',     # Green
        'negative': 'EF4444',     # Red
        'neutral': '888888',      # Gray
        'light_bg': 'F8F9FA',     # Light gray
        'white': 'FFFFFF',
    }
    
    def __init__(self, client_config: ClientConfig):
        """Initialize exporter with client configuration."""
        self.config = client_config
        self.wb = None
        self._setup_styles()
    
    def _setup_styles(self):
        """Create named styles for consistent formatting."""
        self.styles = {
            'header': {
                'font': Font(bold=True, color='FFFFFF', size=11),
                'fill': PatternFill('solid', fgColor=self.COLORS['header']),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    bottom=Side(style='thin', color='000000')
                )
            },
            'subheader': {
                'font': Font(bold=True, color=self.COLORS['primary'], size=10),
                'fill': PatternFill('solid', fgColor=self.COLORS['light_bg']),
            },
            'metric_label': {
                'font': Font(bold=True, size=10),
                'alignment': Alignment(horizontal='left'),
            },
            'metric_value': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='right'),
            },
            'positive': {
                'font': Font(color=self.COLORS['positive'], bold=True),
            },
            'negative': {
                'font': Font(color=self.COLORS['negative'], bold=True),
            },
        }
    
    def _apply_style(self, cell, style_name: str):
        """Apply a named style to a cell."""
        style = self.styles.get(style_name, {})
        for attr, value in style.items():
            setattr(cell, attr, value)
    
    def export(
        self,
        report_data: Dict[str, Any],
        filename: str = None
    ) -> Path:
        """
        Export report data to Excel.
        
        Args:
            report_data: Full report data dictionary
            filename: Output filename (auto-generated if not provided)
        
        Returns:
            Path to created Excel file
        """
        self.wb = Workbook()
        
        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        
        # Create sheets
        self._create_executive_summary(report_data)
        self._create_traffic_overview(report_data)
        self._create_search_performance(report_data)
        self._create_content_performance(report_data)
        self._create_audience_insights(report_data)
        self._create_acquisition_channels(report_data)
        self._create_insights_sheet(report_data)
        
        # Generate filename
        if filename is None:
            period = report_data.get('metadata', {}).get('current_period', {}).get('label', 'report')
            client_name = self.config.name
            filename = f"{client_name}_quarterly_report_{period.replace(' ', '_')}.xlsx"
        
        output_path = OUTPUT_DIR / filename
        self.wb.save(output_path)
        
        return output_path
    
    def _create_executive_summary(self, data: Dict[str, Any]):
        """Create executive summary sheet."""
        ws = self.wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = f"{self.config.display_name} - Quarterly Analytics Report"
        ws['A1'].font = Font(bold=True, size=16, color=self.COLORS['primary'])
        ws.merge_cells('A1:F1')
        
        # Period
        metadata = data.get('metadata', {})
        current = metadata.get('current_period', {})
        previous = metadata.get('previous_period', {})
        
        ws['A3'] = f"Period: {current.get('label', 'N/A')} vs {previous.get('label', 'N/A')}"
        ws['A3'].font = Font(size=12, italic=True)
        
        # Key Metrics Section
        ws['A5'] = "KEY METRICS"
        self._apply_style(ws['A5'], 'header')
        ws.merge_cells('A5:D5')
        
        # Traffic metrics
        traffic = data.get('ga4', {}).get('traffic_overview', {})
        comparison = data.get('comparison', {}).get('traffic_overview', {})
        
        metrics = [
            ('Total Users', 'total_users'),
            ('New Users', 'new_users'),
            ('Sessions', 'sessions'),
            ('Pageviews', 'pageviews'),
            ('Bounce Rate', 'bounce_rate'),
            ('Avg Session Duration (sec)', 'avg_session_duration'),
        ]
        
        row = 7
        ws[f'A{row}'] = 'Metric'
        ws[f'B{row}'] = 'Current'
        ws[f'C{row}'] = 'Previous'
        ws[f'D{row}'] = 'Change'
        
        for col in ['A', 'B', 'C', 'D']:
            self._apply_style(ws[f'{col}{row}'], 'subheader')
        
        row += 1
        for label, key in metrics:
            current_val = traffic.get(key, 0)
            comp_data = comparison.get(key, {})
            prev_val = comp_data.get('previous', 0) if isinstance(comp_data, dict) else 0
            change = comp_data.get('change', {}) if isinstance(comp_data, dict) else {}
            
            ws[f'A{row}'] = label
            ws[f'B{row}'] = current_val
            ws[f'C{row}'] = prev_val
            
            if isinstance(change, dict):
                change_str = change.get('formatted', 'N/A')
                ws[f'D{row}'] = change_str
                if change.get('direction') == 'up':
                    self._apply_style(ws[f'D{row}'], 'positive')
                elif change.get('direction') == 'down':
                    self._apply_style(ws[f'D{row}'], 'negative')
            
            row += 1
        
        # Executive Summary text
        row += 2
        ws[f'A{row}'] = "EXECUTIVE SUMMARY"
        self._apply_style(ws[f'A{row}'], 'header')
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 2
        insights = data.get('insights', {})
        summary = insights.get('executive_summary', 'No summary available.')
        ws[f'A{row}'] = summary
        ws.merge_cells(f'A{row}:F{row + 2}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        # Key Recommendations
        row += 5
        ws[f'A{row}'] = "KEY RECOMMENDATIONS"
        self._apply_style(ws[f'A{row}'], 'header')
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 2
        recommendations = insights.get('key_recommendations', [])
        for i, rec in enumerate(recommendations[:5], 1):
            ws[f'A{row}'] = f"{i}. {rec}"
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def _create_traffic_overview(self, data: Dict[str, Any]):
        """Create traffic overview sheet."""
        ws = self.wb.create_sheet("Traffic Overview")
        
        # Title
        ws['A1'] = "Website Traffic Overview"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['primary'])
        
        # Monthly data
        monthly = data.get('ga4', {}).get('traffic_by_month', pd.DataFrame())
        
        if not monthly.empty:
            ws['A3'] = "Monthly Traffic Breakdown"
            self._apply_style(ws['A3'], 'subheader')
            
            self._write_dataframe(ws, monthly, start_row=5)
    
    def _create_search_performance(self, data: Dict[str, Any]):
        """Create search performance sheet."""
        ws = self.wb.create_sheet("Search Performance")
        
        ws['A1'] = "Google Search Console Performance"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['primary'])
        
        # Overview
        gsc = data.get('gsc', {})
        overview = gsc.get('overview', {})
        
        ws['A3'] = "Search Overview"
        self._apply_style(ws['A3'], 'subheader')
        
        row = 5
        for key, label in [
            ('total_clicks', 'Total Clicks'),
            ('total_impressions', 'Total Impressions'),
            ('avg_ctr', 'Average CTR (%)'),
            ('avg_position', 'Average Position'),
        ]:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = overview.get(key, 0)
            self._apply_style(ws[f'A{row}'], 'metric_label')
            self._apply_style(ws[f'B{row}'], 'metric_value')
            row += 1
        
        # Top keywords
        row += 2
        ws[f'A{row}'] = "Top Keywords by Clicks"
        self._apply_style(ws[f'A{row}'], 'subheader')
        
        keywords = gsc.get('top_keywords_clicks', pd.DataFrame())
        if not keywords.empty:
            self._write_dataframe(ws, keywords.head(20), start_row=row + 2)
        
        # Adjust columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
    
    def _create_content_performance(self, data: Dict[str, Any]):
        """Create content performance sheet."""
        ws = self.wb.create_sheet("Content Performance")
        
        ws['A1'] = "Top Performing Content"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['primary'])
        
        # Top pages
        top_pages = data.get('ga4', {}).get('top_pages', pd.DataFrame())
        
        if not top_pages.empty:
            ws['A3'] = "Top Pages by Pageviews"
            self._apply_style(ws['A3'], 'subheader')
            
            # Select key columns
            display_cols = ['pageTitle', 'pagePath', 'screenPageViews', 'pct_of_total', 
                          'averageSessionDuration', 'bounceRate']
            cols_to_use = [c for c in display_cols if c in top_pages.columns]
            
            self._write_dataframe(ws, top_pages[cols_to_use].head(15), start_row=5)
        
        # Landing pages
        landing = data.get('ga4', {}).get('landing_pages', pd.DataFrame())
        
        if not landing.empty:
            row = 25
            ws[f'A{row}'] = "Top Landing Pages"
            self._apply_style(ws[f'A{row}'], 'subheader')
            
            self._write_dataframe(ws, landing.head(10), start_row=row + 2)
        
        # Adjust columns
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
    
    def _create_audience_insights(self, data: Dict[str, Any]):
        """Create audience insights sheet."""
        ws = self.wb.create_sheet("Audience Insights")
        
        ws['A1'] = "Audience Analysis"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['primary'])
        
        # Device breakdown
        devices = data.get('ga4', {}).get('device_breakdown', pd.DataFrame())
        
        if not devices.empty:
            ws['A3'] = "Traffic by Device"
            self._apply_style(ws['A3'], 'subheader')
            self._write_dataframe(ws, devices, start_row=5)
        
        # Geography
        geo = data.get('ga4', {}).get('geography', pd.DataFrame())
        
        if not geo.empty:
            ws['A12'] = "Traffic by Country"
            self._apply_style(ws['A12'], 'subheader')
            self._write_dataframe(ws, geo.head(10), start_row=14)
        
        # New vs Returning
        nvr = data.get('ga4', {}).get('new_vs_returning', {})
        
        if nvr:
            ws['A28'] = "New vs Returning Visitors"
            self._apply_style(ws['A28'], 'subheader')
            
            row = 30
            for user_type in ['new', 'returning']:
                user_data = nvr.get(user_type, {})
                ws[f'A{row}'] = user_type.title()
                ws[f'B{row}'] = f"{user_data.get('users', 0):,} users"
                ws[f'C{row}'] = f"{user_data.get('pct_of_total', 0)}%"
                row += 1
    
    def _create_acquisition_channels(self, data: Dict[str, Any]):
        """Create acquisition channels sheet."""
        ws = self.wb.create_sheet("Acquisition")
        
        ws['A1'] = "Traffic Acquisition Channels"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['primary'])
        
        # Channel breakdown
        channels = data.get('ga4', {}).get('traffic_by_channel', pd.DataFrame())
        
        if not channels.empty:
            ws['A3'] = "Traffic by Channel"
            self._apply_style(ws['A3'], 'subheader')
            self._write_dataframe(ws, channels, start_row=5)
        
        # Paid search
        paid = data.get('ga4', {}).get('paid_search', {})
        
        if paid and paid.get('sessions', 0) > 0:
            row = len(channels) + 10 if not channels.empty else 15
            ws[f'A{row}'] = "Paid Search Performance"
            self._apply_style(ws[f'A{row}'], 'subheader')
            
            row += 2
            for key, label in [
                ('sessions', 'Sessions'),
                ('users', 'Users'),
                ('bounce_rate', 'Bounce Rate (%)'),
                ('avg_session_duration', 'Avg Duration (sec)'),
            ]:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = paid.get(key, 0)
                row += 1
    
    def _create_insights_sheet(self, data: Dict[str, Any]):
        """Create insights and recommendations sheet."""
        ws = self.wb.create_sheet("Insights")
        
        ws['A1'] = "Analytics Insights & Recommendations"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['primary'])
        
        insights_data = data.get('insights', {})
        insights_list = insights_data.get('insights', [])
        
        if not insights_list:
            ws['A3'] = "No insights generated. Run analysis with data."
            return
        
        # Headers
        row = 3
        headers = ['Priority', 'Category', 'Type', 'Finding', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, 'header')
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 60
        
        # Data rows
        row = 4
        for insight in insights_list:
            ws.cell(row=row, column=1, value=insight.get('priority', ''))
            ws.cell(row=row, column=2, value=insight.get('category', ''))
            ws.cell(row=row, column=3, value=insight.get('type', ''))
            ws.cell(row=row, column=4, value=insight.get('headline', ''))
            ws.cell(row=row, column=5, value=insight.get('recommendation', ''))
            
            # Color by type
            type_col = ws.cell(row=row, column=3)
            if insight.get('type') == 'positive':
                type_col.font = Font(color=self.COLORS['positive'])
            elif insight.get('type') == 'negative':
                type_col.font = Font(color=self.COLORS['negative'])
            elif insight.get('type') == 'opportunity':
                type_col.font = Font(color=self.COLORS['secondary'])
            
            row += 1
    
    def _write_dataframe(
        self,
        ws,
        df: pd.DataFrame,
        start_row: int = 1,
        include_index: bool = False
    ):
        """Write a DataFrame to worksheet with formatting."""
        if df.empty:
            return
        
        # Headers
        col = 1
        if include_index:
            cell = ws.cell(row=start_row, column=col, value=df.index.name or 'Index')
            self._apply_style(cell, 'header')
            col += 1
        
        for header in df.columns:
            cell = ws.cell(row=start_row, column=col, value=header)
            self._apply_style(cell, 'header')
            col += 1
        
        # Data rows
        row = start_row + 1
        for idx, data_row in df.iterrows():
            col = 1
            if include_index:
                ws.cell(row=row, column=col, value=idx)
                col += 1
            
            for value in data_row:
                ws.cell(row=row, column=col, value=value)
                col += 1
            row += 1

